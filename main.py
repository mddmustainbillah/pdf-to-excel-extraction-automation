import base64
import json
import os
import time
import random
import openpyxl
from typing import Dict, Any
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from copy import copy
import google.generativeai as genai
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

###########################################
# PDF Processing Functions
###########################################

def process_pdf(pdf_path, max_retries=5, initial_delay=2):
    """Process PDF directly with Gemini API and return extracted data."""
    
    for attempt in range(max_retries):      
        try:
            # Read and encode PDF
            try:
                with open(pdf_path, 'rb') as file:
                    pdf_bytes = file.read()
                    pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
            except Exception as e:
                print(f"Error reading PDF {pdf_path}: {str(e)}")
                return None

            # Configure Gemini
            genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-flash')

            # Create the prompt
            prompt = """
Extract the following information from the purchase order PDF and return it as JSON (as explained below):
Don't include anything extra that is not in the text.

Your task is to:
1. Extract the following 13 fields from the text (as explained below). If you can't find the field, return null.
2. Translate each extracted value into the Slovak language (except where noted).
3. Return the result as a JSON object, where both keys and structure stay in English, but values are translated into Slovak.
4. Return only the final JSON output — no extra explanation.

Fields:
1. Client: – Only the client name, not the address.
2. Order Number – e.g., From "ORDER2025079" extract "O2025079". Keep O from the "ORDER" (first character) and remove the rest characters(RDER) then keep only the numbers. The add the "O" to the beginning of the number.
3. Foil: – Found under the "SPECIFICATIONS" section and before the "Item name and description" section. Extract the value of the line. Do not convert to slovak.
4. Return of Bulk Containers: – Found under the "SPECIFICATIONS" section and before the "Item name and description" section also start with a dash (`-`). Extract the value of the line. Convert to slovak.
5. Microbiological Analysis: – Found under the "SPECIFICATIONS" section and before the "Item name and description" section also start with a dash (`-`). Extract the value of the line. Convert to slovak.
6. Specific order requirements – Extract information STRICTLY from the "ORDER SPECIFICATIONS" section using these rules:
   - The section starts after "ORDER SPECIFICATIONS:" and ends before "Item name and description"
   - Within this section, collect ALL lines that start with a dash (-) or en dash (–)
   - ONLY exclude these two specific lines (case-insensitive):
     * The line containing "Microbiological analysis:"
     * The line containing "Return of bulk containers:"
   - Keep ALL OTHER lines that start with a dash, including:
     * Archive information
     * Packaging instructions
     * Product handling instructions
     * Any other specifications with a leading dash
   - For each kept line:
     * Remove the leading dash/en dash and trim whitespace
     * Translate the text to Slovak
   - Combine all processed lines with '\n' (literal characters, not actual newlines)

   Example 1:
   Input section:
   ORDER SPECIFICATIONS:
   Foil: PET/AL/LDPE - gloss
   Print: digital CMYK surface print
   - Microbiological analysis: NO
   - Return of bulk containers: YES
   - Archive: YES, 20 pcs mix
   - Mix all products well before packaging.

   Expected output:
   "Archív: ÁNO, 20 ks mix\nZmiešajte všetky produkty pred balením."

   Example 2:
   Input section:
   ORDER SPECIFICATIONS:
   Foil: PET/AL/LDPE - gloss
   Print: digital CMYK surface print
   - Microbiological analysis: NO
   - Return of bulk containers: YES
   - Embossing data: WW = sachets production week
   - Archive: YES, 20 pcs mix

   Expected output:
   "Embossingové údaje: WW = týždeň výroby sáčkov\nArchív: ÁNO, 20 ks mix"

    
    
Important Rule for All Fields(7-13):
    Only extract values that are found **within each product's row** in the **Item name and description** section of the table.  
    Ignore any text found in headers, column titles, footers, or any content outside of the actual product rows.

    Each product section starts with a product name and continues until either:
    - The next product name is found, OR
    - The end of the product details is found
    A single product's information may span across multiple pages. Always collect ALL information for a product across pages before moving to the next product

    
7. Item Name – Extract the product name following these rules:
   - Extract the complete product name from the first line of each product section
   - Keep the name in its original language (Slovak, English, or other) - DO NOT translate
   - Include the Art. number/Art. no. if present (e.g., "Art.: 277137")
   - Extract everything before any product ID numbers
   
   Examples:
   Input: "1. HYALURONIC ACTIVE+ CREAM RICH, Art.: 277137"
   Correct output: "HYALURONIC ACTIVE+ CREAM RICH, Art.: 277137"
   
   Input: "2. Anglický humor"
   Correct output: "Anglický humor"

   Keep in mind do not include product id (e.g. MINIPAK product ID: O2025097 - 1) in the item name. Always Item Name will be before product id.

8. Sachet Size – Find the line starting with `Sachet size / filling volume:` and extract only the dimensions (e.g., `60x100`). Ignore units like "mm".
    Correct pattern: Sachet size / filling volume: 60x100mm* / 10ml (+/-0,2ml)
    Incorrect pattern: Filling volume: 10ml (+/- 0,2ml)
    Incorrect pattern: Sachet size: 60x100mm
9. Filling Volume – From the same line of Sachet Size, extract the value after the slash `/`. If you find direct "Filling volume: 10ml (+/- 0,2ml)" then skip this field as it is not matching the pattern.
    Rules:
    - If the value is like "10ml", return "10ml".
    - If the value is like "3,5ml", return "3.5ml" (replace comma with dot).
    - If the value is like "3ml (+/-0,2ml)", return "3ml (+/-0.2ml)".
    - If the value is like "3,5ml (+/-0,2ml)", return "3.5ml (+/-0.2ml)".

10. Products Heating – Extract the value of the line. (Translate to Slovak)
11. Embossing Data – Extract the value of the line. (Translate to Slovak if has no number)
12. Required Bulk Quantity – Amount with unit. Example: 25kg, if there is 25,5kg then extract 25.5kg.
13. Qty – Total quantity (multiply quantity by unit, e.g., "25 1000 pcs" = "25000", "7,5 100 pcs" = "750"). Replace the comma with a dot.

Any values found outside the actual product rows — like in the column headers, footnotes, or lead time section — must be ignored for all fields above.

Return the data in this format:
{
  "Client Name": "...",
  "Order Number": "...",
  "Foil": "...",
  "Return of Bulk Containers": "...",
  "Microbiological Analysis": "...",
  "Specific order requirements": "...",
  "Items": [
    {
      "Item Name": "...",
      "Sachet Size": "...",
      "Filling Volume": "...",
      "Products Heating": "...",
      "Embossing Data": "...",
      "Required Bulk Quantity": "...",
      "Qty": "..."
    },
    {
      "Item Name": "...",
      "Sachet Size": "...",
      "Filling Volume": "...",
      "Products Heating": "...",
      "Embossing Data": "...",
      "Required Bulk Quantity": "...",
      "Qty": "..."
    }
    // ... more items if present
  ]
}

Ensure the output is a valid JSON object. Do not include any additional text or explanations outside of the JSON structure.
"""
            
            # Create PDF part object
            pdf_part = {
                "mime_type": "application/pdf",
                "data": pdf_base64
            }
            
            # Generate content with both prompt and PDF
            response = model.generate_content(
                [prompt, pdf_part],
                generation_config={
                    'temperature': 0.1,
                    'top_p': 0.1,
                    'top_k': 16,
                }
            )
            
            # Process response
            response_text = response.text.strip()
            
            try:
                # Clean response text
                response_text = response_text.strip('```json').strip('```').strip()
                
                # Parse JSON and return directly
                return json.loads(response_text)
                
            except json.JSONDecodeError as e:
                print(f"JSON parsing error: {str(e)}")
                if attempt < max_retries - 1:
                    continue
                return None
                
        except Exception as e:
            if "429" in str(e):
                wait_time = (initial_delay * (2 ** attempt) + random.random())
                print(f"Rate limit exceeded. Retrying in {wait_time:.2f} seconds (attempt {attempt + 1}/{max_retries})")
                time.sleep(wait_time)
            else:
                print(f"Error processing PDF: {str(e)}")
                if attempt < max_retries - 1:
                    continue
                return None
    
    print(f"Failed to process PDF after {max_retries} retries.")
    return None

###########################################
# Excel Processing Functions
###########################################

def load_excel_template(template_file: str) -> openpyxl.Workbook:
    """Load Excel template"""
    return openpyxl.load_workbook(template_file)

def safely_unmerge_row_cells(ws, row: int):
    """Safely unmerge any merged cells in the specified row"""
    try:
        ranges_to_unmerge = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row:
                ranges_to_unmerge.append(str(merged_range))
        
        for range_str in ranges_to_unmerge:
            try:
                ws.unmerge_cells(range_str)
            except (ValueError, KeyError):
                continue
    except Exception:
        # If any error occurs during unmerging, just continue
        pass

def copy_row_format(ws, source_row: int, target_row: int):
    """Copy row format including height and cell styles"""
    try:
        # Copy row height
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

        # First unmerge any cells in the target row
        safely_unmerge_row_cells(ws, target_row)

        # Copy cell formats
        for col in range(1, 13):  # Columns A through L
            try:
                source = ws.cell(row=source_row, column=col)
                target = ws.cell(row=target_row, column=col)
                
                # Copy cell format
                if source.has_style:
                    target.font = copy(source.font)
                    target.border = copy(source.border)
                    target.fill = copy(source.fill)
                    target.number_format = copy(source.number_format)
                    target.protection = copy(source.protection)
                    target.alignment = copy(source.alignment)
            except Exception:
                continue

        # Special handling for the red "Fakturovať:" cell
        last_cell = ws.cell(row=target_row, column=12)  # Column L
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        last_cell.fill = red_fill
    except Exception as e:
        print(f"Warning: Error in copying row format: {str(e)}")

def prepare_product_rows(ws, num_items: int):
    """Prepare rows for all products, inserting new ones if needed"""
    if num_items <= 2:
        return  # Template already has 2 rows
        
    # We need to add (num_items - 2) rows after row 13
    for i in range(num_items - 2):
        try:
            # Insert row at position 14 + i
            ws.insert_rows(14 + i)
            # Copy format from row 13 (second product row) to new row
            copy_row_format(ws, 13, 14 + i)
            # Set the row number in the first column
            ws.cell(row=14 + i, column=1, value=f"{3 + i}.")
        except Exception as e:
            print(f"Warning: Error in preparing row {14 + i}: {str(e)}")

def _recreate_expedicia_section(ws, new_start_row: int):
    """Recreates the Expedícia objednávky header at the new position with hardcoded formatting."""
    
    # Define the exact header range
    header_range_str = f'I{new_start_row}:L{new_start_row}'
    
    # 1. Explicitly unmerge the header range if it exists as a merged cell
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == header_range_str:
            try:
                ws.unmerge_cells(header_range_str)
            except (ValueError, KeyError):
                pass # Already unmerged or not a valid range
            break # Found and unmerged, exit loop

    # 2. Clear content and formatting of individual cells within the header range
    for col_idx in range(9, 13): # Columns I (9) to L (12)
        cell = ws.cell(row=new_start_row, column=col_idx)
        cell.value = None
        cell.font = Font()
        cell.fill = PatternFill()
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = 'General'
        cell.protection = None

    # Define common border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 3. Merge cells for the header
    ws.merge_cells(header_range_str)
    
    # 4. Get the top-left cell of the merged range (which is the only writable cell)
    header_cell = ws.cell(row=new_start_row, column=9)  # Column I
    
    # 5. Set value, alignment, font, and border for the merged cell
    header_cell.value = "Expedícia objednávky"
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.font = Font(bold=True)
    header_cell.border = thin_border

    # Set a fixed row height for the header to ensure vertical centering
    ws.row_dimensions[new_start_row].height = 20 # Adjust as needed for proper display

def set_product_data(ws, row: int, product: Dict[str, Any]):
    """Set product data in the correct cells with proper formatting"""
    try:
        # Ensure no merged cells in this row
        safely_unmerge_row_cells(ws, row)
        
        # Basic product information
        ws.cell(row=row, column=1, value=f"{row-11}.")  # Row number
        ws.cell(row=row, column=2, value=product.get('Item Name', ''))
        ws.cell(row=row, column=3, value=product.get('Embossing Data', ''))
        ws.cell(row=row, column=5, value=product.get('Products Heating', ''))
        ws.cell(row=row, column=6, value=product.get('Sachet Size', ''))
        ws.cell(row=row, column=7, value=product.get('Filling Volume', ''))
        ws.cell(row=row, column=8, value=product.get('Qty', ''))
        ws.cell(row=row, column=9, value=product.get('Required Bulk Quantity', ''))
        ws.cell(row=row, column=10, value='/')
    except Exception as e:
        print(f"Warning: Error in setting product data for row {row}: {str(e)}")

def map_data_to_excel(wb: openpyxl.Workbook, data: Dict[str, Any]) -> openpyxl.Workbook:
    """Map JSON data to Excel template"""
    ws = wb.active
    
    try:
        # Basic information mapping
        ws['C3'] = data.get('Client Name', '')  # (1)
        ws['C4'] = data.get('Order Number', '')  # (2)
        ws['C5'] = data.get('Foil', '')  # (3)
        ws['C6'] = data.get('Return of Bulk Containers', '')  # (4)
        ws['C7'] = data.get('Microbiological Analysis', '')  # (5)
        
        # Handle Specific order requirements in the info box (I4:L9)
        specific_reqs = data.get('Specific order requirements', '')
        if specific_reqs:  # Only split if there's content
            specific_reqs = specific_reqs.split('\n')
        else:
            specific_reqs = ['']  # Empty list with one empty string to maintain structure
        
        # Explicitly unmerge I4:L9 if it's already merged
        for merged_range in list(ws.merged_cells.ranges):
            if str(merged_range) == 'I4:L9':
                try:
                    ws.unmerge_cells('I4:L9')
                except (ValueError, KeyError):
                    pass # Already unmerged or not a valid range
                break

        # Merge cells I4:L9 to create the info box
        ws.merge_cells('I4:L9')
        
        # Configure the merged cell
        cell = ws['I4']  # The main cell of the merged range
        cell.value = '\n'.join(specific_reqs)  # Join all lines with newline
        
        # Set alignment and text wrapping
        cell.alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrapText=True
        )
        
        # Adjust row heights to accommodate text
        for row in range(4, 10):  # Rows 4 to 9
            ws.row_dimensions[row].height = 20
        
        # Handle multiple products
        items = data.get('Items', [])
        prepare_product_rows(ws, len(items))
        
        # Map all products
        for idx, product in enumerate(items, start=1):
            row = 11 + idx
            set_product_data(ws, row, product)
        
        # Recreate Expedícia section at its new position
        # It should be 2 rows after the last product row
        last_product_row = 11 + len(items)
        new_expedicia_start_row = last_product_row + 2
        _recreate_expedicia_section(ws, new_expedicia_start_row)
        
        return wb
    except Exception as e:
        print(f"Warning: Error in mapping data: {str(e)}")
        return wb

def process_excel_imputation(data: Dict[str, Any], template_file: str, output_file: str):
    """Process the Excel imputation directly from data"""
    try:
        # Load template
        wb = load_excel_template(template_file)
        
        # Map data and save
        wb = map_data_to_excel(wb, data)
        wb.save(output_file)
        
        print(f"Successfully created Excel file: {output_file}")
    except Exception as e:
        print(f"Error processing Excel imputation: {str(e)}")

def process_all_pdfs(input_folder: str, template_file: str, output_folder: str):
    """Process all PDFs in the input folder and create Excel files in the output folder"""
    
    # Create output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)
    
    # Get all PDF files from input folder
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        print(f"No PDF files found in {input_folder}")
        return
    
    print(f"Found {len(pdf_files)} PDF files to process")
    
    # Process each PDF file
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(input_folder, pdf_file)
        
        try:
            print(f"\nProcessing file {index} of {len(pdf_files)}: {pdf_file}")
            
            # Step 1: Process PDF and get JSON data
            print("Step 1: Processing PDF and extracting data...")
            extracted_data = process_pdf(pdf_path)
            
            if not extracted_data:
                print(f"Failed to extract data from {pdf_file}. Skipping to next file.")
                continue
            
            # Create output Excel file name (preserve original name structure)
            excel_filename = os.path.splitext(pdf_file)[0] + '_filled_order_note.xlsx'
            output_excel = os.path.join(output_folder, excel_filename)
            
            # Step 2: Process Excel imputation directly with extracted data
            print("Step 2: Creating Excel file with extracted data...")
            process_excel_imputation(extracted_data, template_file, output_excel)
            
            print(f"Successfully processed {pdf_file}")
            print(f"Excel file created: {excel_filename}")
            time.sleep(10) # Add a delay between processing each PDF to respect API rate limits
            
        except Exception as e:
            print(f"Error processing {pdf_file}: {str(e)}")
            continue
    
    print("\nAll files processed!")

def main():
    # Define folder paths
    input_folder = "input_pdfs"  # Folder containing PDF files
    template_folder = "files"    # Folder containing the template
    output_folder = "output_excel"  # Folder where Excel files will be saved
    
    # Template file path
    template_file = os.path.join(template_folder, "empty file for extraction excel file.xlsx")
    
    # Check if input folder exists
    if not os.path.exists(input_folder):
        print(f"Input folder '{input_folder}' does not exist. Creating it...")
        os.makedirs(input_folder)
        print(f"Please place your PDF files in the '{input_folder}' folder and run the script again.")
        return
    
    # Check if template file exists
    if not os.path.exists(template_file):
        print(f"Template file not found at: {template_file}")
        return
    
    try:
        # Process all PDFs
        process_all_pdfs(input_folder, template_file, output_folder)
        
    except Exception as e:
        print(f"Error in main process: {str(e)}")

if __name__ == "__main__":
    main()
