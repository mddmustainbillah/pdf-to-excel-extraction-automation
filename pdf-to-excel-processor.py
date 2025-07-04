import base64
import json
import os
import time
import random
import openpyxl
from typing import Dict, Any
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from copy import copy
import google.generativeai as genai
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

###########################################
# PDF Processing Functions
###########################################

def process_pdf(pdf_path, max_retries=5, initial_delay=1):
    """Process PDF directly with Gemini API and return extracted data."""
    
    for attempt in range(max_retries):      
        try:
            # Read and encode PDF
            try:
                with open(pdf_path, 'rb') as file:
                    pdf_bytes = file.read()
                    pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
            except Exception as e:
                print(f"Error reading PDF {pdf_path}: {str(e)}")
                return None

            # Configure Gemini
            genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
            model = genai.GenerativeModel('gemini-1.5-flash')

            # Create the prompt
            prompt = """
Extract the following information from the purchase order PDF and return it as JSON (as explained below):
Don't include anything extra that is not in the text.

Your task is to:
1. Extract the following 13 fields from the text (as explained below). If you can't find the field, return null.
2. Translate each extracted value into the Slovak language (except where noted).
3. Return the result as a JSON object, where both keys and structure stay in English, but values are translated into Slovak.
4. Return only the final JSON output — no extra explanation.

Fields:
1. Client: – Only the client name, not the address.
2. Order Number – e.g., From "ORDER2025079" extract "O2025079". Keep O from the "ORDER" (first character) and remove the rest characters(RDER) then keep only the numbers. The add the "O" to the beginning of the number.
3. Foil: – Found under the "SPECIFICATIONS" section and before the "Item name and description" section. Extract the value of the line. Do not convert to slovak.
4. Return of Bulk Containers: – Found under the "SPECIFICATIONS" section and before the "Item name and description" section also start with a dash (`-`). Extract the value of the line. Convert to slovak.
5. Microbiological Analysis: – Found under the "SPECIFICATIONS" section and before the "Item name and description" section also start with a dash (`-`). Extract the value of the line. Convert to slovak.
6. Specific order requirements – Extract information STRICTLY from the "ORDER SPECIFICATIONS" section using these rules:
   - The section starts after "ORDER SPECIFICATIONS:" and ends before "Item name and description"
   - Within this section, collect ALL lines that start with a dash (-) or en dash (–)
   - ONLY exclude these two specific lines (case-insensitive):
     * The line containing "Microbiological analysis:"
     * The line containing "Return of bulk containers:"
   - Keep ALL OTHER lines that start with a dash, including:
     * Archive information
     * Packaging instructions
     * Product handling instructions
     * Any other specifications with a leading dash
   - For each kept line:
     * Remove the leading dash/en dash and trim whitespace
     * Translate the text to Slovak
   - Combine all processed lines with '\n' (literal characters, not actual newlines)

   Example 1:
   Input section:
   ORDER SPECIFICATIONS:
   Foil: PET/AL/LDPE - gloss
   Print: digital CMYK surface print
   - Microbiological analysis: NO
   - Return of bulk containers: YES
   - Archive: YES, 20 pcs mix
   - Mix all products well before packaging.

   Expected output:
   "Archív: ÁNO, 20 ks mix\nZmiešajte všetky produkty pred balením."

   Example 2:
   Input section:
   ORDER SPECIFICATIONS:
   Foil: PET/AL/LDPE - gloss
   Print: digital CMYK surface print
   - Microbiological analysis: NO
   - Return of bulk containers: YES
   - Embossing data: WW = sachets production week
   - Archive: YES, 20 pcs mix

   Expected output:
   "Embossingové údaje: WW = týždeň výroby sáčkov\nArchív: ÁNO, 20 ks mix"

    
    
Important Rule for All Fields(7-13):
    Only extract values that are found **within each product's row** in the **Item name and description** section of the table.  
    Ignore any text found in headers, column titles, footers, or any content outside of the actual product rows.

    Each product section starts with a product name and continues until either:
    - The next product name is found, OR
    - The end of the product details is found
    A single product's information may span across multiple pages. Always collect ALL information for a product across pages before moving to the next product
    Maintain the order of products as they appear in the document.

    
7. Item Name – Extract the product name following these rules:
   - Extract the complete product name from the first line of each product section
   - Keep the name in its original language (Slovak, English, or other) - DO NOT translate
   - Include the Art. number/Art. no. if present (e.g., "Art.: 277137")
   - Extract everything before any product ID numbers
   
   Examples:
   Input 1: "HYALURONIC ACTIVE+ CREAM RICH, Art.: 277137"
   Correct output: "HYALURONIC ACTIVE+ CREAM RICH, Art.: 277137"
   
   Input 2: "Anglický humor"
   Correct output: "Anglický humor"

   Input 3: "Age Decode"
    Correct output: "Age Decode"

   Keep in mind do not include product id (e.g. MINIPAK product ID: O2025097 - 1) in the item name. Always Item Name will be before product id.

8. Sachet Size – Find the line starting with `Sachet size / filling volume:` and extract only the dimensions (e.g., `60x100`). Ignore units like "mm".
    Correct pattern: Sachet size / filling volume: 60x100mm* / 10ml (+/-0,2ml)
    Incorrect pattern: Filling volume: 10ml (+/- 0,2ml)
    Incorrect pattern: Sachet size: 60x100mm
9. Filling Volume – From the same line of Sachet Size, extract the value after the slash `/`. If you find direct "Filling volume: 10ml (+/- 0,2ml)" then skip this field as it is not matching the pattern.
    Rules:
    - If the value is like "10ml", return "10ml".
    - If the value is like "3,5ml", return "3.5ml" (replace comma with dot).
    - If the value is like "3ml (+/-0,2ml)", return "3ml (+/-0.2ml)".
    - If the value is like "3,5ml (+/-0,2ml)", return "3.5ml (+/-0.2ml)".

10. Products Heating – Extract the value of the line. (Translate to Slovak)
11. Embossing Data – Extract the value of the line. (Translate to Slovak if has no number)
12. Required Bulk Quantity – Amount with unit. Example: 25kg, if there is 25,5kg then extract 25.5kg.
13. Qty – Total quantity (multiply quantity by unit, e.g., "25 1000 pcs" = "25000", "7,5 100 pcs" = "750"). Replace the comma with a dot.

Any values found outside the actual product rows — like in the column headers, footnotes, or lead time section — must be ignored for all fields above.

Return the data in this format:
{
  "Client Name": "...",
  "Order Number": "...",
  "Foil": "...",
  "Return of Bulk Containers": "...",
  "Microbiological Analysis": "...",
  "Specific order requirements": "...",
  "Items": [
    {
      "Item Name": "...",
      "Sachet Size": "...",
      "Filling Volume": "...",
      "Products Heating": "...",
      "Embossing Data": "...",
      "Required Bulk Quantity": "...",
      "Qty": "..."
    },
    {
      "Item Name": "...",
      "Sachet Size": "...",
      "Filling Volume": "...",
      "Products Heating": "...",
      "Embossing Data": "...",
      "Required Bulk Quantity": "...",
      "Qty": "..."
    }
    // ... more items if present
  ]
}

Ensure the output is a valid JSON object. Do not include any additional text or explanations outside of the JSON structure.
"""
            
            # Create PDF part object
            pdf_part = {
                "mime_type": "application/pdf",
                "data": pdf_base64
            }
            
            # Generate content with both prompt and PDF
            response = model.generate_content(
                [prompt, pdf_part],
                generation_config={
                    'temperature': 0.1,
                    'top_p': 0.1,
                    'top_k': 16,
                }
            )
            
            # Process response
            response_text = response.text.strip()
            
            try:
                # Clean response text
                response_text = response_text.strip('```json').strip('```').strip()
                
                # Parse JSON and return directly
                return json.loads(response_text)
                
            except json.JSONDecodeError as e:
                print(f"JSON parsing error: {str(e)}")
                if attempt < max_retries - 1:
                    continue
                return None
                
        except Exception as e:
            if "429" in str(e):
                wait_time = (initial_delay * (2 ** attempt) + random.random())
                print(f"Rate limit exceeded. Retrying in {wait_time:.2f} seconds (attempt {attempt + 1}/{max_retries})")
                time.sleep(wait_time)
            else:
                print(f"Error processing PDF: {str(e)}")
                if attempt < max_retries - 1:
                    continue
                return None
    
    print(f"Failed to process PDF after {max_retries} retries.")
    return None

###########################################
# Excel Processing Functions
###########################################

def load_excel_template(template_file: str) -> openpyxl.Workbook:
    """Load Excel template"""
    return openpyxl.load_workbook(template_file)

def safely_unmerge_row_cells(ws, row: int):
    """Safely unmerge any merged cells in the specified row - IMPROVED VERSION"""
    try:
        ranges_to_unmerge = []
        # Create a copy of merged_cells to avoid modification during iteration
        merged_ranges = list(ws.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            if merged_range.min_row <= row <= merged_range.max_row:
                ranges_to_unmerge.append(merged_range)
        
        for merged_range in ranges_to_unmerge:
            try:
                ws.unmerge_cells(str(merged_range))
                print(f"Unmerged range: {merged_range}")
            except (ValueError, KeyError) as e:
                print(f"Could not unmerge range {merged_range}: {e}")
                continue
    except Exception as e:
        print(f"Error in safely_unmerge_row_cells: {e}")
        pass

def copy_row_format(ws, source_row: int, target_row: int):
    """Copy row format including height and cell styles"""
    try:
        # Copy row height
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

        # First unmerge any cells in the target row
        safely_unmerge_row_cells(ws, target_row)

        # Copy cell formats
        for col in range(1, 13):  # Columns A through L
            try:
                source = ws.cell(row=source_row, column=col)
                target = ws.cell(row=target_row, column=col)
                
                # Copy cell format
                if source.has_style:
                    target.font = copy(source.font)
                    target.border = copy(source.border)
                    target.fill = copy(source.fill)
                    target.number_format = copy(source.number_format)
                    target.protection = copy(source.protection)
                    target.alignment = copy(source.alignment)
            except Exception:
                continue

        # Special handling for the red "Fakturovať:" cell
        last_cell = ws.cell(row=target_row, column=12)  # Column L
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        last_cell.fill = red_fill
    except Exception as e:
        print(f"Warning: Error in copying row format: {str(e)}")

def prepare_product_rows(ws, num_items: int):
    """Prepare rows for all products, inserting new ones if needed"""
    if num_items <= 2:
        return  # Template already has 2 rows
        
    # We need to add (num_items - 2) rows after row 13
    for i in range(num_items - 2):
        try:
            # Insert row
            ws.insert_rows(14 + i)
            # Copy format from row 13 (second product row) to new row
            copy_row_format(ws, 13, 14 + i)
            # Set the row number in the first column
            ws.cell(row=14 + i, column=1, value=f"{3 + i}.")
        except Exception as e:
            print(f"Warning: Error in preparing row {14 + i}: {str(e)}")

def set_product_data(ws, row: int, product: Dict[str, Any]):
    """Set product data in the correct cells with proper formatting"""
    try:
        # Ensure no merged cells in this row
        safely_unmerge_row_cells(ws, row)
        
        # Basic product information
        ws.cell(row=row, column=1, value=f"{row-11}.")  # Row number
        ws.cell(row=row, column=2, value=product.get('Item Name', ''))
        ws.cell(row=row, column=3, value=product.get('Embossing Data', ''))
        ws.cell(row=row, column=5, value=product.get('Products Heating', ''))
        ws.cell(row=row, column=6, value=product.get('Sachet Size', ''))
        ws.cell(row=row, column=7, value=product.get('Filling Volume', ''))
        ws.cell(row=row, column=8, value=product.get('Qty', ''))
        ws.cell(row=row, column=9, value=product.get('Required Bulk Quantity', ''))
        ws.cell(row=row, column=10, value='/')
    except Exception as e:
        print(f"Warning: Error in setting product data for row {row}: {str(e)}")

def add_additional_rows(ws, last_item_row: int):
    """Add 9 additional rows after the last item with specific formatting - FIXED VERSION"""
    try:
        start_row = last_item_row + 1
        print(f"Adding 9 additional rows starting from row {start_row}")

        # Add 9 rows
        for i in range(9):
            current_row = start_row + i

            # First, check if we need to unmerge any cells that might interfere
            safely_unmerge_row_cells(ws, current_row)

            # Insert row if needed (this ensures the row exists)
            if current_row > ws.max_row:
                ws.insert_rows(current_row)

            # Set row height to 25px
            ws.row_dimensions[current_row].height = 25

            # Add content to columns safely (avoiding merged cells)
            try:
                # Column A (start margin)
                start_cell = ws.cell(row=current_row, column=1)
                # Only set value if it's not a merged cell
                if not isinstance(start_cell, openpyxl.cell.MergedCell):
                    start_cell.value = ""

                # Column L (end margin)
                end_cell = ws.cell(row=current_row, column=12)
                # Only set value if it's not a merged cell
                if not isinstance(end_cell, openpyxl.cell.MergedCell):
                    end_cell.value = ""

                # Clear any existing content in the middle columns (B-K)
                for col in range(2, 12):  # Columns B through K
                    cell = ws.cell(row=current_row, column=col)
                    # Only modify if it's not a merged cell
                    if not isinstance(cell, openpyxl.cell.MergedCell):
                        cell.value = ""
                        cell.alignment = Alignment(horizontal='left', vertical='center')

            except Exception as cell_error:
                print(f"Warning: Error setting cell content in row {current_row}: {cell_error}")
                continue

        print(f"Successfully added 8 additional rows starting from row {start_row}")
        return start_row + 8  # Return the last added row number

    except Exception as e:
        print(f"Error in add_additional_rows: {e}")
        return last_item_row  # Return original last row if failed
    
def format_expedition_row(ws, last_item_row: int):
    """
    Formats the second newly added row for 'Expedícia objednávky'.

    This function merges columns I, J, K, and L in the second row after the last item,
    adds the specified text, and applies bolding, font size, alignment, and a border.

    Args:
        ws: The openpyxl worksheet object.
        last_item_row: The row number of the last item, before adding new rows.
    """
    try:
        # The 8 new rows start at last_item_row + 1.
        # We need to work on the second of these rows.
        target_row = last_item_row + 2

        print(f"Formatting expedition row at actual row number {target_row}")

        # Define the cell range to merge: columns I, J, K, L
        merge_start_col = 9  # Column I
        merge_end_col = 12   # Column L

        # Merge the cells first
        ws.merge_cells(
            start_row=target_row,
            start_column=merge_start_col,
            end_row=target_row,
            end_column=merge_end_col
        )

        # Get the top-left cell of the merged range to apply value and core styles
        merged_cell = ws.cell(row=target_row, column=merge_start_col)

        # 1. Set the value and main styles on the top-left cell
        merged_cell.value = 'Expedícia objednávky'
        merged_cell.font = Font(name='Calibri', size=20, bold=True)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 2. Define the border style
        side = Side(style='medium', color='000000')
        border_style = Border(left=side, right=side, top=side, bottom=side)

        # 3. Apply the border to all cells within the merged range
        for col in range(merge_start_col, merge_end_col + 1):
            cell_to_style = ws.cell(row=target_row, column=col)
            cell_to_style.border = border_style

        print(f"Successfully formatted 'Expedícia objednávky' in merged cells at row {target_row}")

    except Exception as e:
        print(f"An error occurred in format_expedition_row: {e}")


def format_palette_info_row(ws, last_item_row: int):
    """
    Formats the third newly added row for palette information.

    This function formats the row for 'Typ palety', 'Rozmer', and 'Váha'
    by setting values, merging cells, and applying borders.

    Args:
        ws: The openpyxl worksheet object.
        last_item_row: The row number of the last item, before adding new rows.
    """
    try:
        # The 'Expedícia objednávky' row is at last_item_row + 2.
        # This function targets the next row, which is the third new row.
        target_row = last_item_row + 3

        print(f"Formatting palette info row at actual row number {target_row}")

        # Define a reusable thin border style
        side = Side(style='thin', color='000000')
        border_style = Border(left=side, right=side, bottom=side)

        # Define a reusable alignment style
        alignment_style = Alignment(horizontal='left', vertical='center')

        # Task 1: Column I - 'Typ palty'
        cell_i = ws.cell(row=target_row, column=9)
        cell_i.value = 'Typ palty'
        cell_i.border = border_style
        cell_i.alignment = alignment_style
        cell_i.font = Font(name='Calibri', size=16)

        # Task 2: Merge columns J and K for 'Rozmer'
        ws.merge_cells(start_row=target_row, start_column=10, end_row=target_row, end_column=11)
        cell_j = ws.cell(row=target_row, column=10)
        cell_j.value = 'Rozmer'
        cell_j.alignment = alignment_style
        cell_j.font = Font(name='Calibri', size=16)

        # Apply border to both cells in the merged range to ensure it draws correctly
        ws.cell(row=target_row, column=10).border = border_style
        ws.cell(row=target_row, column=11).border = border_style

        # Task 3: Column L - 'Váha' (assuming L, as K is merged)
        cell_l = ws.cell(row=target_row, column=12)
        cell_l.value = 'Váha'
        cell_l.border = border_style
        cell_l.alignment = alignment_style
        cell_l.font = Font(name='Calibri', size=16)

        print(f"Successfully formatted palette info row at row {target_row}")

    except Exception as e:
        print(f"An error occurred in format_palette_info_row: {e}")


def format_final_rows(ws, last_item_row: int):
    """
    Formats the final 4 newly added rows (rows 4, 5, 6, and 7).

    This function performs two main tasks:
    1. Adds a left-side border to each cell in column I for these four rows.
    2. Merges the 8 cells across columns J and K for these four rows into a
       single large box and applies a border to it.

    Args:
        ws: The openpyxl worksheet object.
        last_item_row: The row number of the last item, before adding new rows.
    """
    try:
        # The target rows are the 4th, 5th, 6th, and 7th added rows.
        start_row = last_item_row + 4
        end_row = last_item_row + 7

        print(f"Formatting final notes section from row {start_row} to {end_row}")

        # Define the border styles we will need
        left_border = Border(left=Side(style='thin', color='000000'))

        full_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Task 1: Add a left border to column I for each of the four rows
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=9)  # Column I
            cell.border = left_border

        # Task 2: Merge columns J and K across all four rows into one large box
        ws.merge_cells(
            start_row=start_row,
            start_column=10,  # Column J
            end_row=end_row,
            end_column=11     # Column K
        )

        # Apply a border to all cells within the merged J-K range.
        for row in range(start_row, end_row + 1):
            for col in range(10, 12):  # Columns J and K
                ws.cell(row=row, column=col).border = full_border

        print(f"Successfully formatted the final 4 rows.")

    except Exception as e:
        print(f"An error occurred in format_final_rows: {e}")

def format_footer_rows(ws, last_item_row: int):
    """
    Formats the final footer rows (the 8th and 9th added rows).

    This function assumes the rows have already been created and performs two tasks:
    1. On the 8th row after the items, it adds a left border to columns I, J, K, and L.
    2. On the 9th row, it applies a full border to all cells (A–L) to create a
       final, fully enclosed row.

    Args:
        ws: The openpyxl worksheet object.
        last_item_row: The row number of the last item, before adding new rows.
    """
    try:
        # Define the row numbers for the last two rows.
        signature_row_num = last_item_row + 8
        final_border_row_num = last_item_row + 9

        print(f"Formatting footer rows at {signature_row_num} and {final_border_row_num}")

        # Task 1: Format the 8th added row
        # Apply a left border to columns I, J, K, and L.
        left_border = Border(left=Side(style='thin', color='000000'))
        for col in range(9, 13):  # Columns I through L
            cell = ws.cell(row=signature_row_num, column=col)
            cell.border = left_border

        # Task 2: Apply a full border to all cells in the 9th added row
        full_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        for col in range(1, 13):  # Columns A through L
            cell = ws.cell(row=final_border_row_num, column=col)
            cell.border = full_border

        print("Successfully formatted the footer rows.")

    except Exception as e:
        print(f"An error occurred in format_footer_rows: {e}")

def map_data_to_excel(wb: openpyxl.Workbook, data: Dict[str, Any]) -> openpyxl.Workbook:
    """Map JSON data to Excel template - FIXED VERSION"""
    ws = wb.active
    
    try:
        # Basic information mapping
        ws['C3'] = data.get('Client Name', '')  # (1)
        ws['C4'] = data.get('Order Number', '')  # (2)
        ws['C5'] = data.get('Foil', '')  # (3)
        ws['C6'] = data.get('Return of Bulk Containers', '')  # (4)
        ws['C7'] = data.get('Microbiological Analysis', '')  # (5)
        
        # Handle Specific order requirements in the info box (I4:L9)
        specific_reqs = data.get('Specific order requirements', '')
        if specific_reqs:  # Only split if there's content
            specific_reqs = specific_reqs.split('\n')
        else:
            specific_reqs = ['']  # Empty list with one empty string to maintain structure
        
        # Merge cells I4:L9 to create the info box
        ws.merge_cells('I4:L9')
        
        # Configure the merged cell
        cell = ws['I4']  # The main cell of the merged range
        cell.value = '\n'.join(specific_reqs)  # Join all lines with newline
        
        # Set alignment and text wrapping
        cell.alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrapText=True
        )
        
        # Adjust row heights to accommodate text
        for row in range(4, 10):  # Rows 4 to 9
            ws.row_dimensions[row].height = 20
        
        # Handle multiple products
        items = data.get('Items', [])
        prepare_product_rows(ws, len(items))
        
        # Map all products
        for idx, product in enumerate(items, start=1):
            row = 11 + idx
            set_product_data(ws, row, product)
        
        # Calculate last item row
        last_item_row = 11 + len(items) if items else 12
        
        print(f"Last item row: {last_item_row}")
        
        # Clean up any existing extra rows BEFORE adding new ones
        current_max_row = ws.max_row
        if current_max_row > last_item_row:
            print(f"Cleaning extra rows from {last_item_row + 1} to {current_max_row}")
            
            # Unmerge cells in rows we're about to delete
            for row_to_clean in range(last_item_row + 1, current_max_row + 1):
                safely_unmerge_row_cells(ws, row_to_clean)
            
            # Delete the extra rows
            rows_to_delete = current_max_row - last_item_row
            if rows_to_delete > 0:
                try:
                    ws.delete_rows(last_item_row + 1, rows_to_delete)
                    print(f"Deleted {rows_to_delete} extra rows")
                except Exception as e:
                    print(f"Could not delete rows: {e}")
        
        # NOW add the 7 additional rows (after cleaning)
        final_row = add_additional_rows(ws, last_item_row)
        print(f"Final row after adding additional rows: {final_row}")

        # Format the expedition row
        format_expedition_row(ws, last_item_row)

        format_palette_info_row(ws, last_item_row)

        format_final_rows(ws, last_item_row)

        format_footer_rows(ws, last_item_row)

    

        
        return wb
        
    except Exception as e:
        print(f"Warning: Error in mapping data: {str(e)}")
        return wb

def process_excel_imputation(data: Dict[str, Any], template_file: str, output_file: str):
    """Process the Excel imputation directly from data"""
    try:
        # Load template
        wb = load_excel_template(template_file)
        
        # Map data
        wb = map_data_to_excel(wb, data)
        
        # Save the file
        wb.save(output_file)
        
        print(f"Successfully created Excel file: {output_file}")
    except Exception as e:
        print(f"Error processing Excel imputation: {str(e)}")

def process_all_pdfs(input_folder: str, template_file: str, output_folder: str):
    """Process all PDFs in the input folder and create Excel files in the output folder"""
    
    # Create output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)
    
    # Get all PDF files from input folder
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        print(f"No PDF files found in {input_folder}")
        return
    
    print(f"Found {len(pdf_files)} PDF files to process")
    
    # Process each PDF file
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(input_folder, pdf_file)
        
        try:
            print(f"\nProcessing file {index} of {len(pdf_files)}: {pdf_file}")
            
            # Step 1: Process PDF and get JSON data
            print("Step 1: Processing PDF and extracting data...")
            extracted_data = process_pdf(pdf_path)
            
            if not extracted_data:
                print(f"Failed to extract data from {pdf_file}. Skipping to next file.")
                continue
            
            # Create output Excel file name (preserve original name structure)
            excel_filename = os.path.splitext(pdf_file)[0] + '_filled_order_note.xlsx'
            output_excel = os.path.join(output_folder, excel_filename)
            
            # Step 2: Process Excel imputation directly with extracted data
            print("Step 2: Creating Excel file with extracted data...")
            process_excel_imputation(extracted_data, template_file, output_excel)
            
            print(f"Successfully processed {pdf_file}")
            print(f"Excel file created: {excel_filename}")
            
        except Exception as e:
            print(f"Error processing {pdf_file}: {str(e)}")
            continue
    
    print("\nAll files processed!")

def main():
    # Define folder paths
    input_folder = "input_pdfs"  # Folder containing PDF files
    template_folder = "files"    # Folder containing the template
    output_folder = "output_excel"  # Folder where Excel files will be saved
    
    # Template file path
    template_file = os.path.join(template_folder, "empty file for extraction excel file.xlsx")
    
    # Check if input folder exists
    if not os.path.exists(input_folder):
        print(f"Input folder '{input_folder}' does not exist. Creating it...")
        os.makedirs(input_folder)
        print(f"Please place your PDF files in the '{input_folder}' folder and run the script again.")
        return
    
    # Check if template file exists
    if not os.path.exists(template_file):
        print(f"Template file not found at: {template_file}")
        return
    
    try:
        # Process all PDFs
        process_all_pdfs(input_folder, template_file, output_folder)
        
    except Exception as e:
        print(f"Error in main process: {str(e)}")

if __name__ == "__main__":
    main()